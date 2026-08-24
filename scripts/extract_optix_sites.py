#!/usr/bin/env python3
"""Extract validated Optix site coordinates as compact, map-ready JSON."""

from __future__ import annotations

import argparse
import json
import math
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


ROMANIA_BOUNDS = {
    "minimum_latitude": 43.5,
    "maximum_latitude": 48.4,
    "minimum_longitude": 20.0,
    "maximum_longitude": 30.0,
}


def clean(value: object) -> str:
    return " ".join(str(value or "").split())


def parse_decimal(value: object) -> float | None:
    if value is None:
        return None
    try:
        coordinate = float(str(value).strip().replace(",", "."))
    except ValueError:
        return None
    return coordinate if math.isfinite(coordinate) else None


def is_globally_valid(latitude: float, longitude: float) -> bool:
    return (
        -90 <= latitude <= 90
        and -180 <= longitude <= 180
        and not (latitude == 0 and longitude == 0)
    )


def is_in_romania(latitude: float, longitude: float) -> bool:
    return (
        ROMANIA_BOUNDS["minimum_latitude"] <= latitude <= ROMANIA_BOUNDS["maximum_latitude"]
        and ROMANIA_BOUNDS["minimum_longitude"] <= longitude <= ROMANIA_BOUNDS["maximum_longitude"]
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("input", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()

    workbook = load_workbook(args.input, read_only=True, data_only=True)
    worksheet = workbook.active
    expected_headers = [
        "Region",
        "Site Name",
        "Latitude Decimal",
        "Longitude Decimal",
        "Description",
    ]
    actual_headers = [clean(cell.value) for cell in next(worksheet.iter_rows(max_row=1))]
    if actual_headers[:5] != expected_headers:
        raise ValueError(f"Unexpected workbook headers: {actual_headers[:5]}")

    records: list[list[object]] = []
    rejected = Counter()

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        region, site_name, raw_latitude, raw_longitude, description, *_ = row
        clean_region = clean(region)
        clean_site_name = clean(site_name)
        if not clean_site_name:
            rejected["missing_site_name"] += 1
            continue

        latitude = parse_decimal(raw_latitude)
        longitude = parse_decimal(raw_longitude)
        if latitude is None or longitude is None:
            rejected["missing_coordinates"] += 1
            continue
        if not is_globally_valid(latitude, longitude):
            rejected["invalid_coordinates"] += 1
            continue

        is_international = clean_region.lower().startswith("international")
        if not is_international and not is_in_romania(latitude, longitude):
            rejected["outside_expected_area"] += 1
            continue

        # Compact schema: [site name/code, description, region, latitude, longitude].
        records.append(
            [
                clean_site_name,
                clean(description),
                clean_region,
                round(latitude, 6),
                round(longitude, 6),
            ]
        )

    records.sort(key=lambda record: (record[3], record[4], record[0]))
    payload = {
        "source": args.input.name,
        "valid": len(records),
        "rejected": dict(rejected),
        "schema": ["code", "description", "region", "lat", "lon"],
        "sites": records,
    }

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    print(
        f"Exported {len(records)} valid sites to {args.output}; "
        f"rejected {sum(rejected.values())}: {dict(rejected)}"
    )


if __name__ == "__main__":
    main()
